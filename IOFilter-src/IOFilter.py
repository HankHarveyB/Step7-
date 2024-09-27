import re
import openpyxl
import openpyxl.styles 
import html
from urllib import parse
import requests
import Translator

# 定义边框样式
thin_border = openpyxl.styles.Border(
    left=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
    top=openpyxl.styles.Side(style='thin'),
    bottom=openpyxl.styles.Side(style='thin')
)
# 中文翻译成英文并将空格替换成_
translator = Translator.Translator(from_lang="zh", to_lang="en")
def translate(SourceText):
    return translator.translate(SourceText).replace(' ', '_')
# def translate_and_replace(sourceText):
#     translator = Translator(service_urls=[
#       'translate.google.com'])
#     translated = translator.translate(sourceText, dest='en').text
#     result = translated.replace(' ', '_')
#     return result

# GOOGLE_TRANSLATE_URL = 'http://translate.google.com/m?q=%s&tl=%s&sl=%s'

# def translate(text, to_language="en", text_language="zh"):

#     text = parse.quote(text)
#     url = GOOGLE_TRANSLATE_URL % (text,to_language,text_language)
#     response = requests.get(url)
#     data = response.text
#     expr = r'(?s)class="(?:t0|result-container)">(.*?)<'
#     result = re.findall(expr, data)
#     if (len(result) == 0):
#         return ""

#     return html.unescape(result[0]).replace(' ', '_')



# 读取Excel文件
excel_file = 'cs9动作表已整理已排序(1) - 副本.xlsx'
wb = openpyxl.load_workbook(excel_file)

# 读取符号表汇总
summary_sheet = wb['符号表汇总']
symbol_dict = {}
for row in summary_sheet.iter_rows(min_row=1, max_col=4, values_only=True):
    symbol_dict[re.sub(r'\s+', ' ',  row[1])] = (row[0], row[3])  # 使用地址作为键，符号和注释作为值

# 读取STL源文件
with open('CS9FbFc.AWL', 'r', encoding='gb2312') as f:
    stl_content = f.read()

# 定义正则表达式，匹配I、Q、PIW、PQW地址
address_pattern = re.compile(r'\b(I|Q|PIW|PQW)\s*([\d\.]+)')

NoneStrCounter=1;
STLCode="";
StarDBNum=300;
DBNumStep=1;
# 遍历设备sheet
for sheet in wb.sheetnames[4:]:  # 从第五个sheet开始是设备
    
    STLCode+=f"""
DATA_BLOCK DB {StarDBNum}
TITLE = {sheet}
VERSION : 0.1


  STRUCT 	

    """
    
    

    device_sheet = wb[sheet]
    print("find sheet:"+sheet)
    
    # 从第一行的E列开始，遍历直到找到第一个空列
    blocks = [cell.value for cell in device_sheet[1][4:] if cell.value]

    # # 读取E列的FC/FB块编号
    # blocks = [cell.value for cell in device_sheet['E'] if cell.value]
    
    # 遍历每一个块编号，查找STL文件中对应的代码段
    for block in blocks:
        print("------------find block:"+block)
        block_pattern = re.compile(rf'(FUNCTION|FUNCTION_BLOCK) ({block[0:2]}) ({block[2:]}).*?END_FUNCTION', re.DOTALL)
        block_match = block_pattern.search(stl_content)
        print("--------------------------find stl:")
        print(block_match)
        if block_match:
            block_code = block_match.group()
            # print("-----------------------------------block code is"+block_code)
            # 在块代码中查找I、Q、PIW、PQW地址
            addresses = address_pattern.findall(block_code)
            
            addresses = list(set(addresses))

            
            for addr_type, addr_value in addresses:
                addr = f"{addr_type} {addr_value}"
                addr = re.sub(r'\s+', ' ', addr)  # 去除多余空格
                print("------------------------------------addr is"+addr)
                
                 # 检查 device_sheet 中是否已经包含了该地址
                addr_exists = False
                for row in device_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0] == addr:  # 如果第一列已经包含该地址
                        addr_exists = True
                        break

                if not addr_exists:
                    # 查找符号表中的符号和注释
                    if addr in symbol_dict:
                        # 将地址、符号、注释写入设备对应的sheet
                        ShouldTranslate = True#是否联网进行翻译
                        symbol, comment = symbol_dict[addr]
                        symbol_en=Translator.process_string(translate(symbol)) if symbol is not None and ShouldTranslate else "None"
                        symbol=symbol if symbol is not None else "None"
                        comment=comment if comment is not None else "None"

                        print("----------------------------------------------------symbol and comment is "+symbol_en+"-"+comment)
                        device_sheet.append([addr, symbol, comment,symbol_en])
                        
                        #TestBOOL : BOOL ;	//临时占位符变量1
                        symbol_en_short=Translator.string_lengthLimit(symbol_en.replace('-','')).replace('.','').replace('/','').replace('\\','').replace('(','').replace(')','')
                        if(symbol_en_short=="None"):
                            symbol_en_short=f"NoSymbol_{NoneStrCounter}"
                            NoneStrCounter+=1
                        if(addr.startswith('I') or addr.startswith('Q')):
                            STLCode+=f"\n   {symbol_en_short} : BOOL ;	//{addr} {symbol}({symbol_en})"
                        elif(addr.startswith('PIW') or addr.startswith('PQW')):
                            STLCode+=f"\n   {symbol_en_short} : WORD ;	//{addr} {symbol}({symbol_en})"
                        else:
                            pass
                            

                        # row = device_sheet.max_row + 1  # 找到最后一行，之后开始写入
                        # device_sheet[f'A{row}'] = addr
                        # device_sheet[f'B{row}'] = symbol if symbol is not None else "None"
                        # device_sheet[f'C{row}'] = comment if comment is not None else " None"
                    else:
                        print("----------------------------------------------------this addr has no symbol!")
                        device_sheet.append([addr])
                        

                        
                        symbol_en=f"NoSymbol_{NoneStrCounter}"
                        NoneStrCounter+=1
                        if(addr.startswith('I') or addr.startswith('Q')):
                            STLCode+=f"\n   {symbol_en} : BOOL ;	//{addr} {symbol_en}"
                        elif(addr.startswith('PIW') or addr.startswith('PQW')):
                            STLCode+=f"\n   {symbol_en} : WORD ;	//{addr} {symbol_en}"
                        else:
                            pass
                        # row = device_sheet.max_row + 1  # 找到最后一行，之后开始写入
                        # device_sheet[f'A{row}'] = addr

                    

                    # 获取刚刚插入的数据的行号
                    new_row = device_sheet.max_row

                    # 设置新插入行的单元格边框
                    for col in ['A', 'B', 'C','D']:  # A列到C列
                        cell = device_sheet[f'{col}{new_row}']
                        cell.border = thin_border   
                        

     # 获取数据部分（不包括第一行表头）
    data = []
    for row in device_sheet.iter_rows(min_row=2, max_row=device_sheet.max_row, max_col=4, values_only=True):
        data.append(row)
    
    # 按第一列（地址）进行升序排序
    sorted_data = sorted(data, key=lambda x: x[0])  # 按地址列排序
    
    # 将排序后的数据重新写回表格
    for i, row in enumerate(sorted_data, start=2):  # 从第二行开始重新写入
        for j, value in enumerate(row, start=1):  # A列到C列
            device_sheet.cell(row=i, column=j, value=value)
    STLCode+="""\n  END_STRUCT ;	
BEGIN
END_DATA_BLOCK

    """
    StarDBNum+=DBNumStep   
    
#print(translate("工业自动化"))
    
# 保存更新后的Excel文件
wb.save('CS9updated_file.xlsx')
print("updated_file.xlsx update!")



with open("CS9DBSTL.awl", 'w', encoding='gb2312') as file:
    file.write(STLCode)
print("DBSTL.awl update!")
